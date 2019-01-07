import csv
import click
from openpyxl import load_workbook

@click.command()
@click.argument('inputfile', nargs=1)
@click.argument('outputfile', nargs=1)
def main(inputfile, outputfile):
    # Loading Excel file
    wb = load_workbook(inputfile)
    ws = wb.active

    # Assigning regions of data
    plate1L = ws['B5':'G12']
    plate1R = ws['H5':'M12']

    plate2L = ws['B18':'G25']
    plate2R = ws['H18':'M25']

    barcode1a = [ws['O6'].value]
    barcode1b = [ws['O7'].value]
    barcode1c = [ws['P6'].value]
    barcode1d = [ws['P7'].value]

    barcode2a = [ws['O19'].value]
    barcode2b = [ws['O20'].value]
    barcode2c = [ws['P19'].value]
    barcode2d = [ws['P20'].value]

    # Defining columns and populating with static data
    sample_id = []
    sentrix_barcode = barcode1a * 24 + barcode1b * 24 + barcode1c * 24 + barcode1d * 24 + \
                    barcode2a * 24 + barcode2b * 24 + barcode2c * 24 + barcode2d * 24

    sentrix_pos = ['R01C01','R03C01','R05C01','R07C01','R09C01','R11C01','R02C01','R04C01','R06C01','R08C01','R10C01','R12C01',
                'R01C02','R03C02','R05C02','R07C02','R09C02','R11C02','R02C02','R04C02','R06C02','R08C02','R10C02','R12C02']
    sentrix_pos *= 8


    sample_well = ['A01','A02','A03','A04','A05','A06','B01','B02','B03','B04','B05','B06','C01','C02','C03','C04','C05','C06','D01','D02','D03','D04','D05','D06',
                'E01','E02','E03','E04','E05','E06','F01','F02','F03','F04','F05','F06','G01','G02','G03','G04','G05','G06','H01','H02','H03','H04','H05','H06',
                'A07','A08','A09','A10','A11','A12','B07','B08','B09','B10','B11','B12','C07','C08','C09','C10','C11','C12','D07','D08','D09','D10','D11','D12',
                'E07','E08','E09','E10','E11','E12','F07','F08','F09','F10','F11','F12','G07','G08','G09','G10','G11','G12','H07','H08','H09','H10','H11','H12']
    sample_well *= 2



    # Populating with data from spreadsheet
    for row in plate1L:
        for cell in row:
            sample_id.append(cell.value)
            
    for row in plate1R:
        for cell in row:
            sample_id.append(cell.value)
            
    for row in plate2L:
        for cell in row:
            sample_id.append(cell.value)
            
    for row in plate2R:
        for cell in row:
            sample_id.append(cell.value)


    #Creating CSV
    with open(outputfile, 'w', newline='') as f:
        csvwriter = csv.writer(f)
        csvwriter.writerow(['[Header]','','','','','','','','','','','','',''])
        csvwriter.writerow(['Institute Name','CAMH','','','','','','','','','','','',''])
        csvwriter.writerow(['Investigator Name','Natalie Freeman','','','','','','','','','','','',''])
        csvwriter.writerow(['Project Name','','','','','','','','','','','','',''])
        csvwriter.writerow(['Date','','','','','','','','','','','','',''])
        csvwriter.writerow(['','','','','','','','','','','','','',''])
        csvwriter.writerow(['[Manifests]','','','','','','','','','','','','',''])
        csvwriter.writerow(['','','','','','','','','','','','','',''])
        csvwriter.writerow(['','','','','','','','','','','','','',''])
        csvwriter.writerow(['[Data]','','','','','','','','','','','','',''])
        csvwriter.writerow(['Sample_ID','SentrixBarcode_A','SentrixPosition_A','Sample_Plate','Sample_Well','Sample_Group','Gender','Sample_Name','Replicate','Parent1','Parent2','','',''])
        for sample, barcode, position, well in zip(sample_id, sentrix_barcode, sentrix_pos, sample_well):
            csvwriter.writerow([sample, barcode, position, '', well, '', '', '', '', '', '', '', '', ''])
        
    print('DONE')

if __name__ == '__main__':
    main()