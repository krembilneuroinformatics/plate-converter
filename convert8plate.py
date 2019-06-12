import csv
import click
from openpyxl import load_workbook

@click.command()
@click.argument('inputfile', nargs=1)
@click.argument('outputfile', nargs=1)
def main(inputfile, outputfile):
    # Loading Excel file
    wb = load_workbook(inputfile)
    ws = wb['Sheet1']
    ws2 = wb['Sheet2']

    # Assigning regions of data
    plate = ws['B5':'I12']
    plate = tuple(zip(*plate))

    barcode1 = [ws['B15'].value]
    barcode2 = [ws['C15'].value]
    barcode3 = [ws['D15'].value]
    barcode4 = [ws['E15'].value]

    barcode5 = [ws['F15'].value]
    barcode6 = [ws['G15'].value]
    barcode7 = [ws['H15'].value]
    barcode8 = [ws['I15'].value]

    project_code = (ws['B17'].value + '_') if ws['B17'].value else ''

    # Read Gender dictionary from Sheet2
    gender_dict = {}
    for i in ws2.values:
        gender_dict[i[0]] = i[1]
    gender_dict.pop('LabID', None)

    # Defining columns and populating with static data
    sample_ids = []
    genders = []
    sentrix_barcodes = barcode1 * 8 + barcode2 * 8 + barcode3 * 8 + barcode4 * 8 + \
                    barcode5 * 8 + barcode6 * 8 + barcode7 * 8 + barcode8 * 8

    sentrix_pos = ['R01C01','R02C01','R03C01','R04C01','R05C01','R06C01','R07C01','R08C01']
    sentrix_pos *= 8

    sample_well = ['A01','B01','C01','D01','E01','F01','G01','H01','A02','B02','C02','D02','E02','F02','G02','H02',
                'A03','B03','C03','D03','E03','F03','G03','H03','A04','B04','C04','D04','E04','F04','G04','H04',
                'A05','B05','C05','D05','E05','F05','G05','H05','A06','B06','C06','D06','E06','F06','G06','H06',
                'A07','B07','C07','D07','E07','F07','G07','H07','A08','B08','C08','D08','E08','F08','G08','H08']


    # Populating with data from spreadsheet
    for column in plate:
        for cell in column:
            sample_ids.append(project_code + str(cell.value))
            genders.append(gender_dict[cell.value])


    #Creating CSV
    with open(outputfile, 'w', newline='') as f:
        csvwriter = csv.writer(f)
        csvwriter.writerow(['[Header]','','','','','','','','','','',''])
        csvwriter.writerow(['Institute Name','CAMH','','','','','','','','','',''])
        csvwriter.writerow(['Investigator Name','Natalie Freeman','','','','','','','','','',''])
        csvwriter.writerow(['Project Name','','','','','','','','','','',''])
        csvwriter.writerow(['Date','','','','','','','','','','',''])
        csvwriter.writerow(['','','','','','','','','','','',''])
        csvwriter.writerow(['[Manifests]','','','','','','','','','','',''])
        csvwriter.writerow(['','','','','','','','','','','',''])
        csvwriter.writerow(['','','','','','','','','','','',''])
        csvwriter.writerow(['[Data]','','','','','','','','','','',''])
        csvwriter.writerow(['Sample_ID','SentrixBarcode_A','SentrixPosition_A','Sample_Plate','Sample_Well','Sample_Group','Gender', 'Sample_Name','Replicate','Parent1','Parent2','2.5Exome_demo'])
        for sample, barcode, position, well, gender in zip(sample_ids, sentrix_barcodes, sentrix_pos, sample_well, genders):
            csvwriter.writerow([sample, barcode, position, '', well, '', gender, '', '', '', '', ''])
        
    print('DONE')

if __name__ == '__main__':
    main()